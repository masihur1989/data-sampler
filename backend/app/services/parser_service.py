"""
Excel Parser Service Module

This module provides memory-efficient Excel file parsing for large files (up to 150MB).
It uses openpyxl's read_only mode for .xlsx files and pandas with chunking for .xls files
to minimize memory usage when processing large datasets.

Key Features:
- Streaming row iteration with configurable chunk sizes
- Support for both .xlsx (openpyxl) and .xls (xlrd) formats
- Lazy loading of sheet names, columns, and row counts
- Memory-efficient processing without loading entire file into memory
"""

from pathlib import Path
from typing import Iterator, Optional
import pandas as pd
from openpyxl import load_workbook

from app.config import EXCEL_CHUNK_SIZE


class ExcelParser:
    """
    Memory-efficient Excel file parser supporting both .xlsx and .xls formats.
    
    This parser is designed to handle large Excel files (up to 150MB) by:
    - Using openpyxl's read_only mode which streams data instead of loading all at once
    - Iterating rows in configurable chunks to limit memory usage
    - Caching metadata (sheet names, columns, row count) to avoid repeated file reads
    
    For .xlsx files: Uses openpyxl with read_only=True for streaming
    For .xls files: Uses xlrd engine with pandas chunking
    """
    
    def __init__(self, file_path: Path):
        """
        Initialize the parser with a file path.
        
        Args:
            file_path: Path to the Excel file (.xlsx or .xls)
        """
        self.file_path = file_path
        self.extension = file_path.suffix.lower()
        self._workbook = None
        self._sheet_names: list[str] = []
        self._columns: list[str] = []
        self._row_count: int = 0

    def get_sheet_names(self) -> list[str]:
        """
        Get list of sheet names in the Excel file.
        Results are cached after first call.
        
        Returns:
            List of sheet names
        """
        if self._sheet_names:
            return self._sheet_names

        if self.extension == ".xlsx":
            # Use read_only mode to minimize memory usage
            wb = load_workbook(self.file_path, read_only=True, data_only=True)
            self._sheet_names = wb.sheetnames
            wb.close()
        else:
            # For .xls files, read just the sheet names without loading data
            df = pd.read_excel(self.file_path, sheet_name=None, nrows=0, engine="xlrd")
            self._sheet_names = list(df.keys())

        return self._sheet_names

    def get_columns(self, sheet_name: Optional[str] = None) -> list[str]:
        """
        Get column names from the first row of the specified sheet.
        
        Args:
            sheet_name: Optional sheet name (uses active sheet if not specified)
            
        Returns:
            List of column names
        """
        if self.extension == ".xlsx":
            wb = load_workbook(self.file_path, read_only=True, data_only=True)
            ws = wb[sheet_name] if sheet_name else wb.active
            first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
            wb.close()
            if first_row:
                # Handle None values in header row
                self._columns = [str(c) if c is not None else f"Column_{i}" for i, c in enumerate(first_row)]
        else:
            df = pd.read_excel(self.file_path, sheet_name=sheet_name or 0, nrows=0, engine="xlrd")
            self._columns = list(df.columns)

        return self._columns

    def get_row_count(self, sheet_name: Optional[str] = None) -> int:
        """
        Get the number of data rows (excluding header) in the specified sheet.
        
        Args:
            sheet_name: Optional sheet name (uses active sheet if not specified)
            
        Returns:
            Number of data rows
        """
        if self.extension == ".xlsx":
            wb = load_workbook(self.file_path, read_only=True, data_only=True)
            ws = wb[sheet_name] if sheet_name else wb.active
            # Subtract 1 for header row
            self._row_count = ws.max_row - 1 if ws.max_row else 0
            wb.close()
        else:
            df = pd.read_excel(self.file_path, sheet_name=sheet_name or 0, engine="xlrd")
            self._row_count = len(df)

        return self._row_count

    def iter_rows(self, sheet_name: Optional[str] = None, chunk_size: int = EXCEL_CHUNK_SIZE) -> Iterator[pd.DataFrame]:
        """
        Iterate over rows in chunks for memory-efficient processing.
        
        This is the primary method for processing large files. It yields
        DataFrames of chunk_size rows at a time, allowing sampling algorithms
        to process data without loading the entire file into memory.
        
        Args:
            sheet_name: Optional sheet name (uses active sheet if not specified)
            chunk_size: Number of rows per chunk (default: 10,000)
            
        Yields:
            DataFrame chunks of the specified size
        """
        if self.extension == ".xlsx":
            yield from self._iter_xlsx_rows(sheet_name, chunk_size)
        else:
            yield from self._iter_xls_rows(sheet_name, chunk_size)

    def _iter_xlsx_rows(self, sheet_name: Optional[str], chunk_size: int) -> Iterator[pd.DataFrame]:
        wb = load_workbook(self.file_path, read_only=True, data_only=True)
        ws = wb[sheet_name] if sheet_name else wb.active

        rows = []
        columns = None
        row_num = 0

        for row in ws.iter_rows(values_only=True):
            if row_num == 0:
                columns = [str(c) if c is not None else f"Column_{i}" for i, c in enumerate(row)]
                row_num += 1
                continue

            rows.append(row)
            row_num += 1

            if len(rows) >= chunk_size:
                yield pd.DataFrame(rows, columns=columns)
                rows = []

        if rows:
            yield pd.DataFrame(rows, columns=columns)

        wb.close()

    def _iter_xls_rows(self, sheet_name: Optional[str], chunk_size: int) -> Iterator[pd.DataFrame]:
        for chunk in pd.read_excel(
            self.file_path,
            sheet_name=sheet_name or 0,
            engine="xlrd",
            chunksize=chunk_size,
        ):
            yield chunk

    def read_all(self, sheet_name: Optional[str] = None) -> pd.DataFrame:
        if self.extension == ".xlsx":
            return pd.read_excel(self.file_path, sheet_name=sheet_name or 0, engine="openpyxl")
        else:
            return pd.read_excel(self.file_path, sheet_name=sheet_name or 0, engine="xlrd")

    def get_file_info(self, sheet_name: Optional[str] = None) -> dict:
        sheets = self.get_sheet_names()
        columns = self.get_columns(sheet_name)
        row_count = self.get_row_count(sheet_name)

        return {
            "sheet_names": sheets,
            "columns": columns,
            "row_count": row_count,
            "column_count": len(columns),
        }
